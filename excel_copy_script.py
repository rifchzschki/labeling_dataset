import openpyxl

# Load the workbook and select the sheet
print('here')
workbook = openpyxl.load_workbook('data_slo_tr_cetak_dan_slo_terbit_oktober_2023.xlsx')
sheet = workbook.active
print('here')

# Define a new workbook to copy data to
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Loop through the rows and copy data
for row in sheet.iter_rows(min_row=1, max_row=200, min_col=1, max_col=200):
    new_row = [cell.value for cell in row]
    new_sheet.append(new_row)

# Save the new workbook
new_workbook.save('slo-excel-data.xlsx')